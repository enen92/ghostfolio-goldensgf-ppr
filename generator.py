#! /usr/bin/env python3
import argparse
from enum import Enum
import json
import os
import unicodedata
import sys
import urllib.request
import time
from openpyxl import load_workbook
from jinja2 import Template

history = {}

GOLDEN_SGF_HISTORY_XSLS = "https://goldensgf.pt/wp-content/uploads/2024/08/HISTORICO-DE-COTACOES.xlsx"
LOCAL_FILE = "history.xlsx"


class OutputFormat(Enum):
    HTML = 1
    JSON = 2
    CSV = 3
    ALL = 4


def download_file(url: str, local_filename: str = LOCAL_FILE):
    with urllib.request.urlopen(url) as response, open(local_filename, 'wb') as out_file:
        file_size = int(response.getheader('Content-Length', 0))
        downloaded = 0
        block_size = 8192  # 8 KB chunks
        start_time = time.time()

        while True:
            buffer = response.read(block_size)
            if not buffer:
                break

            downloaded += len(buffer)
            out_file.write(buffer)

            elapsed_time = time.time() - start_time
            speed = downloaded / elapsed_time / 1024  # in KB/s
            percent = downloaded * 100 / file_size if file_size else 0
            print(
                f"\rDownloaded: {downloaded / 1024:.2f} KB "
                f"({percent:.2f}%) at {speed:.2f} KB/s",
                end=''
            )
    if file_size and downloaded < file_size:
        print(
            f"Download incomplete: expected {file_size} bytes, got {downloaded} bytes")
        return False
    print("\nDownload complete!")
    return True


def grab_history(filename: str = LOCAL_FILE):
    wb = load_workbook(filename)
    for sheet in wb:
        for row in sheet.iter_rows(min_row=2, max_col=3):
            ppr = row[0].value
            value = row[1].value
            date = row[2].value
            if not ppr or not value or not date:
                continue
            if ppr not in history.keys():
                history[ppr] = {
                    "currentMarketPrice": {
                        "value": None,
                        "date": None},
                    "history": []}
            history[ppr]["history"].append((date, value))

    # sort by latest date
    for key, data in history.items():
        sorted(data["history"], key=lambda x: x[1], reverse=True)
        most_recent_timestamp, most_recent_value = data["history"][0]
        history[key]["currentMarketPrice"]["value"] = most_recent_value
        history[key]["currentMarketPrice"]["date"] = most_recent_timestamp


def generate_output(
        ppr_name: str,
        output_format: OutputFormat = OutputFormat.HTML):
    file_name = unicodedata.normalize(
        'NFKD',
        f"{ppr_name.replace(' ', '').lower()}"
    ).encode('ASCII', 'ignore').decode('utf-8')
    ppr_data = history.get(ppr_name, None)
    if output_format == OutputFormat.HTML:
        file_extension = ".html"
        html_template = '''<!DOCTYPE html>
<html>
<head>
    <title>{{title}}</title>
</head>
<style>
body { font-family: Arial; font-size: 0.3cm }
</style>

<body>
    <h1>{{title}}</h1>
    <div id="currentMarketPrice" class="container">
        <h2 class="title">Most Recent Value (UP):</h2>
        <h3 id="currentMarketPrice_upvalue">{{value}}</h3>
        <h3 id="currentMarketPrice_update">{{date}}</h3>
    </div>
    <div id="historyData" class="container">
        <h2 class="title">History data:</h1>
{% for daytime_value in history -%}
{{ daytime_value[0].strftime("%Y-%m-%d") + ";" +  daytime_value[1]|string + "</br>"}}
{% endfor %}
    </div>
</body>
</html>'''
        if ppr_data:
            tm = Template(html_template)
            output = tm.render(
                title=ppr_name,
                value=ppr_data["currentMarketPrice"]["value"],
                date=ppr_data["currentMarketPrice"]["date"].strftime("%Y-%m-%d"),
                history=ppr_data["history"])
            write_output(f"{file_name}{file_extension}", output)
    elif output_format == OutputFormat.JSON:
        file_extension = ".json"
        output = {
            "title": ppr_name,
            "currentMarketPrice": {
                "value": ppr_data["currentMarketPrice"]["value"],
                "date": ppr_data["currentMarketPrice"]["date"].strftime("%Y-%m-%d")},
            "history": [
                (date.strftime("%Y-%m-%d"),
                 value) for date,
                value in ppr_data["history"]]}
        write_output(f"{file_name}{file_extension}", json.dumps(output))
    elif output_format == OutputFormat.CSV:
        file_extension = ".csv"
        output = ["date;marketPrice"]
        output.extend(
            [f"{date.strftime('%Y-%m-%d')};{value}" for date, value in ppr_data["history"]])
        write_output(f"{file_name}{file_extension}", "\n".join(output))


def write_output(file_name: str, payload: str):
    with open(os.path.join("output", file_name), "w") as f:
        f.write(payload)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description='Simple graber for golden SGF PPR historical data. Usage:')
    parser.add_argument(
        '-p',
        '--ppr',
        help='PPR name (e.g. "SGF DR FINANÇAS")',
        required=False)
    parser.add_argument(
        '-l',
        '--list',
        help='List available PPRs',
        action='store_true',
        required=False)
    parser.add_argument(
        '-o',
        '--output',
        help='Output format ("json", "html", "csv"). Use "all" for all.',
        default="all",
        required=False)
    if len(sys.argv) == 1:
        parser.print_help(sys.stderr)
        sys.exit(1)
    args = vars(parser.parse_args())
    if not download_file(GOLDEN_SGF_HISTORY_XSLS):
        sys.exit(1)
    grab_history()
    if args["list"]:
        print("Available options:")
        for key in history.keys():
            print(key)
        sys.exit(0)
    # PPR list (or only selected ones)
    ppr_list = history.keys() if not args["ppr"] else [args["ppr"]]
    # Format
    output_formats = []
    if args["output"] == "json":
        output_formats.append(OutputFormat.JSON)
    elif args["output"] == "csv":
        output_formats.append(OutputFormat.CSV)
    elif args["output"] == "html":
        output_formats.append(OutputFormat.HTML)
    else:
        output_formats = [OutputFormat.JSON, OutputFormat.CSV, OutputFormat.HTML]
    for output_format in output_formats:
        for ppr in ppr_list:
            print(f"Generating for PPR: {ppr} ({output_format})")
            generate_output(ppr, output_format=output_format)
